# file: feral_kitty_fifi/feral_kitty_fifi/features/gimme_report.py
# Python Cog: Roster/Ban/Leave XLSX report with optional backfill from log channels (Railway-friendly).
# Deps (pip): discord.py openpyxl
# Intents needed: Guilds, GuildMembers, GuildMessages, MessageContent, GuildBans
# Permissions: View Channel, Read Message History, Send Messages, Attach Files
# Optional: View Audit Log (improves ban reason)

from __future__ import annotations

import io
import os
import re
import sqlite3
from datetime import datetime, timezone
from typing import Any, Dict, Iterable, List, Optional, Tuple, Union

import discord
from discord.ext import commands
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

# --------------------------
# ENV-FIRST CONFIG (Railway)
# --------------------------
REPORT_CHANNEL_ID = int(os.getenv("REPORT_CHANNEL_ID", "0"))  # required
TRIGGER = os.getenv("TRIGGER", "!gimme")

JOIN_LEAVE_LOG_CHANNEL_IDS: List[int] = [
    int(s) for s in filter(None, (os.getenv("JOIN_LEAVE_LOG_CHANNEL_IDS") or "").split(","))
]  # required (comma-separated)

BAN_LOG_CHANNEL_ID = int(os.getenv("BAN_LOG_CHANNEL_ID", "0"))  # required

BACKFILL_MAX_MESSAGES_PER_CHANNEL = max(
    1, min(200_000, int(os.getenv("BACKFILL_MAX_MESSAGES_PER_CHANNEL", "50000")))
)

DB_PATH = os.getenv(
    "DB_PATH", os.path.join(os.getcwd(), "data", "gimme_report.sqlite")
)

# -------------
# Misc helpers
# -------------
def iso(dt: Optional[datetime | str | float | int]) -> str:
    if dt is None:
        return ""
    if isinstance(dt, datetime):
        d = dt if dt.tzinfo else dt.replace(tzinfo=timezone.utc)
        return d.isoformat()
    # try ISO str
    try:
        d = datetime.fromisoformat(str(dt).replace("Z", "+00:00"))
        if d.tzinfo is None:
            d = d.replace(tzinfo=timezone.utc)
        return d.isoformat()
    except Exception:
        pass
    # try epoch
    try:
        d = datetime.fromtimestamp(float(dt), tz=timezone.utc)
        return d.isoformat()
    except Exception:
        return ""

def days_between(a: str, b: str) -> str:
    try:
        da = datetime.fromisoformat(a.replace("Z", "+00:00"))
        db = datetime.fromisoformat(b.replace("Z", "+00:00"))
        if da.tzinfo is None:
            da = da.replace(tzinfo=timezone.utc)
        if db.tzinfo is None:
            db = db.replace(tzinfo=timezone.utc)
        return str(max(0, (db - da).days))
    except Exception:
        return ""

def normalize_text(s: Optional[str]) -> str:
    return re.sub(r"\s+", " ", (s or "")).strip()

def message_to_searchable_text(m: discord.Message) -> str:
    parts: List[str] = []
    if m.content:
        parts.append(m.content)
    for e in m.embeds or []:
        if e.title:
            parts.append(e.title)
        if e.description:
            parts.append(e.description)
        if e.author and e.author.name:
            parts.append(e.author.name)
        if e.footer and e.footer.text:
            parts.append(e.footer.text)
        for f in getattr(e, "fields", []) or []:
            if getattr(f, "name", None):
                parts.append(f.name)
            if getattr(f, "value", None):
                parts.append(f.value)
    return normalize_text(" | ".join(filter(None, parts)))

def extract_user_id(text: str) -> Optional[int]:
    m1 = re.search(r"<@!?(?P<id>\d{17,20})>", text)
    if m1 and m1.group("id"):
        return int(m1.group("id"))
    m2 = re.search(r"\b(?P<id>\d{17,20})\b", text)
    if m2 and m2.group("id"):
        return int(m2.group("id"))
    return None

def detect_event_type(text: str) -> Optional[str]:
    t = text.lower()
    if re.search(r"\b(banned|ban)\b", t):
        return "ban"
    if re.search(r"\b(left|leave|removed)\b", t):
        return "leave"
    if re.search(r"\b(joined|join)\b", t):
        return "join"
    return None

def extract_reason(text: str) -> Optional[str]:
    m = re.search(r"reason\s*[:\-]\s*(.+)$", text, re.IGNORECASE)
    if m and m.group(1):
        return normalize_text(m.group(1))[:1000]
    return None

# -----------
# SQLite API
# -----------
class DbApi:
    def __init__(self, path: str):
        os.makedirs(os.path.dirname(path), exist_ok=True)
        self.db = sqlite3.connect(path, isolation_level=None)  # autocommit
        self.db.execute("PRAGMA journal_mode = WAL;")
        self._init_schema()

    def _init_schema(self):
        self.db.executescript(
            """
            CREATE TABLE IF NOT EXISTS users (
              user_id TEXT PRIMARY KEY,
              username TEXT NOT NULL,
              discord_created_at TEXT
            );

            CREATE TABLE IF NOT EXISTS events (
              message_id TEXT PRIMARY KEY,
              channel_id TEXT NOT NULL,
              user_id TEXT NOT NULL,
              username TEXT,
              event_type TEXT NOT NULL CHECK(event_type IN ('join','leave','ban')),
              ts TEXT NOT NULL,
              reason TEXT
            );

            CREATE TABLE IF NOT EXISTS meta (
              key TEXT PRIMARY KEY,
              value TEXT
            );

            CREATE INDEX IF NOT EXISTS idx_events_user_ts ON events(user_id, ts);
            CREATE INDEX IF NOT EXISTS idx_events_type_ts ON events(event_type, ts);
            """
        )

    def upsert_user(self, user_id: int | str, username: str, created_at_iso: str) -> None:
        self.db.execute(
            """
            INSERT INTO users (user_id, username, discord_created_at)
            VALUES (?, ?, ?)
            ON CONFLICT(user_id) DO UPDATE SET
              username=excluded.username,
              discord_created_at=COALESCE(excluded.discord_created_at, users.discord_created_at)
            """,
            (str(user_id), username, created_at_iso),
        )

    def insert_event(self, message_id: int | str, channel_id: int | str, user_id: int | str,
                     username: Optional[str], event_type: str, ts_iso: str, reason: Optional[str]) -> int:
        cur = self.db.execute(
            """
            INSERT OR IGNORE INTO events (message_id, channel_id, user_id, username, event_type, ts, reason)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (str(message_id), str(channel_id), str(user_id), username, event_type, ts_iso, reason),
        )
        return cur.rowcount or 0

    def get_meta(self, key: str) -> Optional[str]:
        cur = self.db.execute("SELECT value FROM meta WHERE key=?", (key,))
        row = cur.fetchone()
        return row[0] if row else None

    def set_meta(self, key: str, value: str) -> None:
        self.db.execute(
            """
            INSERT INTO meta (key, value) VALUES (?, ?)
            ON CONFLICT(key) DO UPDATE SET value=excluded.value
            """,
            (key, value),
        )

    def list_users_with_any_event(self) -> List[str]:
        cur = self.db.execute("SELECT DISTINCT user_id FROM events")
        return [r[0] for r in cur.fetchall()]

    def list_events_by_user(self, user_id: str) -> List[Tuple[str, str, str, str, str]]:
        cur = self.db.execute(
            """
            SELECT user_id, COALESCE(username, user_id) AS username, event_type, ts, COALESCE(reason,'')
            FROM events
            WHERE user_id=?
            ORDER BY ts ASC
            """,
            (user_id,),
        )
        return [(r[0], r[1], r[2], r[3], r[4]) for r in cur.fetchall()]

    def close(self):
        try:
            self.db.close()
        except Exception:
            pass

# -------------------------------
# XLSX (openpyxl) report builder
# -------------------------------
def wb_add_header(ws: Worksheet, headers: List[str]):
    ws.append(headers)
    ws.freeze_panes = "A2"

def build_periods(db: DbApi) -> Tuple[List[Dict[str, str]], List[Dict[str, str]]]:
    left_periods: List[Dict[str, str]] = []
    ban_periods: List[Dict[str, str]] = []

    for user_id in db.list_users_with_any_event():
        events = db.list_events_by_user(user_id)
        current_join: Optional[str] = None
        current_username: Optional[str] = None
        for _uid, username, etype, ts, reason in events:
            current_username = current_username or username or user_id
            if etype == "join":
                current_join = ts
                continue
            if etype in ("leave", "ban") and current_join:
                period = {
                    "username": current_username or user_id,
                    "user_id": user_id,
                    "date_joined": current_join,
                    "date_left": ts,
                    "left_type": etype,
                    "reason": reason or "",
                    "total_days": days_between(current_join, ts),
                }
                if etype == "leave":
                    left_periods.append(period)
                else:
                    ban_periods.append(period)
                current_join = None

    left_periods.sort(key=lambda p: p["date_left"], reverse=True)
    ban_periods.sort(key=lambda p: p["date_left"], reverse=True)
    return left_periods, ban_periods

def build_workbook(role_list: List[Tuple[int, str]], members: Iterable[discord.Member], db: DbApi) -> io.BytesIO:
    wb = Workbook()
    # ---------------- Roster
    ws1 = wb.active
    ws1.title = "Roster"
    base_headers = ["discord_created_at", "server_joined_at", "username", "user_id"]
    role_headers = [f"role:{name}" for _rid, name in role_list]
    wb_add_header(ws1, base_headers + role_headers)

    for m in members:
        user = m._user if hasattr(m, "_user") else m.guild.get_member(m.id).user if hasattr(m, "guild") else m  # fallback
        uname = f"{user.name}#{user.discriminator}" if getattr(user, "discriminator", "0") != "0" else user.name
        row = [
            iso(getattr(user, "created_at", None)),
            iso(getattr(m, "joined_at", None)),
            uname,
            str(user.id),
        ]
        member_role_ids = {r.id for r in getattr(m, "roles", [])}
        row.extend([(rid in member_role_ids) for rid, _name in role_list])
        ws1.append(row)

        # Best-effort backfill into DB
        try:
            db.upsert_user(user.id, uname, iso(getattr(user, "created_at", None)))
        except Exception:
            pass

    # ---------------- Bans
    ws2 = wb.create_sheet("Bans")
    wb_add_header(ws2, ["username", "user_id", "date_joined", "date_left", "ban_reason"])

    # ---------------- Left Members
    ws3 = wb.create_sheet("Left Members")
    wb_add_header(ws3, ["username", "user_id", "date_joined", "date_left", "total_days"])

    left_periods, ban_periods = build_periods(db)
    for p in ban_periods:
        ws2.append([p["username"], p["user_id"], p["date_joined"], p["date_left"], p["reason"]])
    for p in left_periods:
        ws3.append([p["username"], p["user_id"], p["date_joined"], p["date_left"], p["total_days"]])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ------------------------
# Backfill implementation
# ------------------------
TextLikeChannel = Union[discord.TextChannel, discord.Thread]

async def fetch_all_messages_chrono(channel: TextLikeChannel, max_messages: int) -> List[discord.Message]:
    """Fetch up to max_messages messages, returning oldest->newest."""
    out: List[discord.Message] = []
    async for m in channel.history(limit=max_messages, oldest_first=True):
        out.append(m)
        if len(out) >= max_messages:
            break
    return out

async def fetch_after_chrono(channel: TextLikeChannel, after_id: int, max_messages: int) -> List[discord.Message]:
    after_obj = discord.Object(id=after_id)
    out: List[discord.Message] = []
    async for m in channel.history(limit=max_messages, after=after_obj, oldest_first=True):
        out.append(m)
        if len(out) >= max_messages:
            break
    return out

def parse_log_message(msg: discord.Message, forced_type: Optional[str] = None) -> Optional[Dict[str, Any]]:
    text = message_to_searchable_text(msg)
    uid = extract_user_id(text)
    if not uid:
        return None
    etype = forced_type or detect_event_type(text)
    if not etype:
        return None
    reason = extract_reason(text) if etype == "ban" else None
    return {
        "message_id": msg.id,
        "channel_id": msg.channel.id,
        "user_id": uid,
        "username": None,
        "event_type": etype,
        "ts": iso(msg.created_at),
        "reason": reason,
    }

async def backfill_from_log_channel(
    guild: discord.Guild,
    channel_id: int,
    forced_type: Optional[str],
    db: DbApi,
    max_messages: int,
) -> Tuple[int, int]:
    channel = guild.get_channel(channel_id)
    if channel is None:
        try:
            channel = await guild.fetch_channel(channel_id)  # type: ignore
        except Exception:
            return (0, 0)

    # Only handle text channels and threads
    if not isinstance(channel, (discord.TextChannel, discord.Thread)):
        return (0, 0)

    meta_key = f"last_message_id:{channel_id}"
    last_id = db.get_meta(meta_key)
    messages: List[discord.Message] = []
    try:
        if last_id:
            messages = await fetch_after_chrono(channel, int(last_id), max_messages)
        else:
            messages = await fetch_all_messages_chrono(channel, max_messages)
    except Exception:
        messages = []

    inserted = 0
    for m in messages:
        evt = parse_log_message(m, forced_type)
        if not evt:
            continue
        try:
            inserted += db.insert_event(
                evt["message_id"],
                evt["channel_id"],
                evt["user_id"],
                evt["username"],
                evt["event_type"],
                evt["ts"],
                evt["reason"],
            )
        except Exception:
            pass

    if messages:
        newest_id = messages[-1].id
        db.set_meta(meta_key, str(newest_id))

    return (len(messages), inserted)

# -----------
# The Cog
# -----------
class GimmeReport(commands.Cog):
    """Generates an XLSX roster report and backfills join/leave/ban events from log channels.

    Trigger: send the exact text defined in TRIGGER (default '!gimme') in any guild text channel.
    """

    def __init__(
        self,
        bot: commands.Bot,
        *,
        db_path: str = DB_PATH,
        trigger: str = TRIGGER,
        report_channel_id: int = REPORT_CHANNEL_ID,
        join_leave_log_channel_ids: Optional[List[int]] = None,
        ban_log_channel_id: int = BAN_LOG_CHANNEL_ID,
        backfill_max_messages_per_channel: int = BACKFILL_MAX_MESSAGES_PER_CHANNEL,
    ):
        self.bot = bot
        self.trigger = trigger
        self.report_channel_id = report_channel_id
        self.join_leave_log_channel_ids = join_leave_log_channel_ids or JOIN_LEAVE_LOG_CHANNEL_IDS
        self.ban_log_channel_id = ban_log_channel_id
        self.backfill_max = backfill_max_messages_per_channel
        self.db = DbApi(db_path)

    def cog_unload(self):
        self.db.close()

    # ----- live tracking (best-effort) -----
    @commands.Cog.listener()
    async def on_member_join(self, member: discord.Member):
        try:
            user = member._user if hasattr(member, "_user") else member.guild.get_member(member.id).user
            uname = f"{user.name}#{user.discriminator}" if getattr(user, "discriminator", "0") != "0" else user.name
            self.db.upsert_user(user.id, uname, iso(getattr(user, "created_at", None)))
            self.db.insert_event(
                f"live:join:{member.guild.id}:{user.id}:{int(discord.utils.utcnow().timestamp()*1000)}",
                "live",
                user.id,
                uname,
                "join",
                iso(getattr(member, "joined_at", discord.utils.utcnow())),
                None,
            )
        except Exception:
            pass

    @commands.Cog.listener()
    async def on_member_remove(self, member: discord.Member):
        try:
            # We may not have the full User object here depending on cache.
            u = getattr(member, "_user", None) or getattr(member, "user", None)
            if isinstance(u, (discord.User, discord.Member)):
                uname = f"{u.name}#{u.discriminator}" if getattr(u, "discriminator", "0") != "0" else u.name
                self.db.upsert_user(u.id, uname, iso(getattr(u, "created_at", None)))
                username = uname
                uid = u.id
            else:
                username = str(member.id)
                uid = member.id
            self.db.insert_event(
                f"live:leave:{member.guild.id}:{member.id}:{int(discord.utils.utcnow().timestamp()*1000)}",
                "live",
                uid,
                username,
                "leave",
                iso(discord.utils.utcnow()),
                None,
            )
        except Exception:
            pass

    @commands.Cog.listener()
    async def on_member_ban(self, guild: discord.Guild, user: discord.User):
        try:
            reason = None
            try:
                # Requires "View Audit Log" permission to enrich reason
                async for entry in guild.audit_logs(limit=6, action=discord.AuditLogAction.ban):
                    if entry.target and entry.target.id == user.id:
                        reason = entry.reason
                        break
            except Exception:
                pass

            uname = f"{user.name}#{user.discriminator}" if getattr(user, "discriminator", "0") != "0" else user.name
            self.db.upsert_user(user.id, uname, iso(getattr(user, "created_at", None)))
            self.db.insert_event(
                f"live:ban:{guild.id}:{user.id}:{int(discord.utils.utcnow().timestamp()*1000)}",
                "live",
                user.id,
                uname,
                "ban",
                iso(discord.utils.utcnow()),
                normalize_text(reason)[:1000] if reason else None,
            )
        except Exception:
            pass

    # ----- command trigger -----
    @commands.Cog.listener()
    async def on_message(self, message: discord.Message):
        try:
            if not message.guild or getattr(message.author, "bot", False):
                return
            if message.content.strip() != self.trigger:
                return

            # Backfill from logs
            for cid in self.join_leave_log_channel_ids:
                try:
                    await backfill_from_log_channel(
                        message.guild, cid, None, self.db, self.backfill_max
                    )
                except Exception:
                    pass

            if self.ban_log_channel_id:
                try:
                    await backfill_from_log_channel(
                        message.guild, self.ban_log_channel_id, "ban", self.db, self.backfill_max
                    )
                except Exception:
                    pass

            # Ensure member cache
            try:
                await message.guild.fetch_members(limit=None).flatten()  # type: ignore[attr-defined]
                members_iter: Iterable[discord.Member] = message.guild.members
            except Exception:
                # Fallback: use cached members only
                members_iter = [m for m in message.guild.members if not getattr(m, "bot", False)]

            # Role list (exclude @everyone), sorted by position desc
            roles_sorted = sorted(
                [r for r in message.guild.roles if not r.is_default()],
                key=lambda r: r.position,
                reverse=True,
            )
            roles_pairs: List[Tuple[int, str]] = []
            seen: Dict[str, int] = {}
            for r in roles_sorted:
                count = seen.get(r.name, 0) + 1
                seen[r.name] = count
                name = f"{r.name} ({count})" if count > 1 else r.name
                roles_pairs.append((r.id, name))

            # Build workbook bytes
            xlsx_bytes = build_workbook(roles_pairs, members_iter, self.db)

            # Resolve report channel
            report_ch = None
            if REPORT_CHANNEL_ID:
                report_ch = message.guild.get_channel(REPORT_CHANNEL_ID)
                if report_ch is None:
                    try:
                        report_ch = await message.guild.fetch_channel(REPORT_CHANNEL_ID)
                    except Exception:
                        report_ch = None
            if not isinstance(report_ch, (discord.TextChannel, discord.Thread)):
                report_ch = message.channel  # fallback to invoking channel

            await report_ch.send(
                content=f"üìä Roster report for **{message.guild.name}**",
                file=discord.File(fp=xlsx_bytes, filename="server_roster.xlsx"),
            )
        except Exception:
            try:
                await message.channel.send("‚ùå Failed to build report.")
            except Exception:
                pass

# -----------
# Extension hook
# -----------
async def setup(bot: commands.Bot):
    bot.add_cog(
        GimmeReport(
            bot,
            db_path=DB_PATH,
            trigger=TRIGGER,
            report_channel_id=REPORT_CHANNEL_ID,
            join_leave_log_channel_ids=JOIN_LEAVE_LOG_CHANNEL_IDS,
            ban_log_channel_id=BAN_LOG_CHANNEL_ID,
            backfill_max_messages_per_channel=BACKFILL_MAX_MESSAGES_PER_CHANNEL,
        )
    )
